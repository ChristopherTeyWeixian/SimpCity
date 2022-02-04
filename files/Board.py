import pathlib
import openpyxl
from openpyxl.cell import cell

class Board(object):
    def __init__(self):
        self.board = []
        self.turn = 1
        self.score = 0
        
        self.Beach= 8
        self.Factory= 8
        self.House= 8
        self.Shop= 8
        self.Highway= 8
        self.Park = 8
        self.Monument = 8

        self.BeachOpt = "True"
        self.FactoryOpt = "True"
        self.HouseOpt = "True"
        self.ShopOpt = "True"
        self.HighwayOpt = "True"
        self.ParkOpt = "False"
        self.MonumentOpt = "False"

        self.BuildingState = [self.BeachOpt, self.FactoryOpt, self.HouseOpt, self.ShopOpt, self.HighwayOpt, self.ParkOpt, self.MonumentOpt]

    def New_Board(self):
        try:
            self.Get_BuildingState()
            # adds_line is needed as "\b" is counted as a functionality in python
            # thus a roundabout way was used instead of just changing the file name
            adds_line = "/q"
            adds_line = adds_line.replace("q", "")

            # current file path not+ \..\data\baseboard.xlxs
            # python files is different location as data files
            path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data" + adds_line + "base_board.xlsx"
    #        print(path)

            # loads the excel file of path
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            # nested for loop to get each cell one by one and append into the board
            for x in range(1, sheet_obj.max_row + 1):
                temp_list = []
                for y in range(1, sheet_obj.max_column + 1):
                    cell_obj = sheet_obj.cell(row = x, column = y).value
                    if cell_obj == "e":
                        cell_obj = cell_obj.replace("e", " ")
                    temp_list.append(cell_obj)
                self.board.append(temp_list)

            return True
        except:
            print("[!] Could not connect to Excel Database, Please Try Again.\n")
            return False

    def Load_Board(self):
        try:
            self.Get_BuildingState()
            # adds_line is needed as "\b" is counted as a functionality in python
            # thus a roundabout way was used instead of just changing the file name
            adds_line = "/q"
            adds_line = adds_line.replace("q", "")

            # current file path not+ \..\data\baseboard.xlxs
            # python files is different location as data files
            path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/save_board.xlsx"
            #print(path)

            # loads the excel file of path
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            # nested for loop to get each cell one by one and append into the board
            for x in range(1, sheet_obj.max_row + 1):
                temp_list = []
                for y in range(1, sheet_obj.max_column + 1):
                    cell_obj = sheet_obj.cell(row = x, column = y).value
                    if cell_obj == "e":
                        cell_obj = cell_obj.replace("e", " ")
                    elif cell_obj == "n":
                        cell_obj = cell_obj.replace("n", "")
                    temp_list.append(cell_obj)
                self.board.append(temp_list)

            path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/save_data.txt"

            txt_data = open(path,"r")
            sv_data = txt_data.readline()
            sv_data = sv_data.split(";")

            self.turn = int(sv_data[0])
            self.Beach = int(sv_data[1])
            self.Factory = int(sv_data[2])
            self.House = int(sv_data[3])
            self.Shop = int(sv_data[4])
            self.Highway = int(sv_data[5])
            self.Park = int(sv_data[6])
            self.Monument = int(sv_data[7])

            txt_data.close()

            return True
        except:
            print("[!] Could not connect to Excel Database, Please Try Again.\n")
            return False

    def Save_Board(self):
        try:
            adds_line = "/q"
            adds_line = adds_line.replace("q", "")
            path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/save_board.xlsx"

            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active

            for x in range(0, len(self.board[0])):
                for y in range(0, len(self.board)):
                    if self.board[y][x] == " ":
                        sheet_obj.cell(row = y + 1, column= x + 1).value = "e"
                    elif self.board[y][x] == "":
                        sheet_obj.cell(row = y + 1, column= x + 1).value = "n"
                    else:
                        sheet_obj.cell(row = y + 1, column= x + 1).value = self.board[y][x]

            wb_obj.save(path)
            
            dataList = [self.turn, self.Beach, self.Factory, self.House, self.Shop, self.Highway, self.Park, self.Monument]

            path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/save_data.txt"

            txt_data = open(path,"w")
            sv_data = ""
            for i in range(0, len(dataList)):
                if i != 0:
                    sv_data = sv_data + ";" + str(dataList[i])
                else:
                    sv_data = sv_data + str(dataList[i])

            txt_data.write(sv_data)
            txt_data.close()
            print("Game saved!")

            return True
        except:
            print("Unable to connect to save file")
            return False

    def Next_Turn(self):
        self.turn += 1
        return self.turn

    def update_board_building(self,Building_Placed):
        if Building_Placed=="BCH":
            self.Beach-=1
            return self.Beach
            
        elif Building_Placed=="FAC":
            self.Factory-=1
            return self.Factory

        elif Building_Placed=="HSE":
            self.House-=1
            return self.House

        elif Building_Placed=="SHP":
            self.Shop-=1
            return self.Shop

        elif Building_Placed=="HWY":
            self.Highway-=1
            return self.Highway

        # 2 new buildings
        elif Building_Placed == "PRK":
            self.Park -= 1
            return self.Park

        elif Building_Placed == "MON":
            self.Monument -= 1
            return self.Monument

    def Change_BuildingState(self, i):
        if i >= len(self.BuildingState) or i < 0:
            return "Out of Range"
        elif i == 0:
            return "Exitting"

        j = i - 1

        amount_Active = 0
        amount_Inactive = 0

        for state in self.BuildingState:
            if state == "True":
                amount_Active += 1
            else:
                amount_Inactive += 1

        if amount_Active == 5 and self.BuildingState[j] == "False":
            print("\n[!] You have already selected 5 buildings !! Please deselect other buildings first\n")
        else:
            if self.BuildingState[j] == "True":
                self.BuildingState[j] = "False"
            else:
                self.BuildingState[j] = "True"
        
        return self.BuildingState[j]

    def Get_BuildingState(self):
        adds_line = "/q"
        adds_line = adds_line.replace("q", "")

        path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/state_data.txt"

        txt_data = open(path,"r")
        sv_data = txt_data.readline()
        sv_data = sv_data.split(";")

        for i in range(0, len(sv_data)):
            self.BuildingState[i] = sv_data[i]

        txt_data.close()
        
        return self.BuildingState

    def Set_BuildingState(self):
        adds_line = "/q"
        adds_line = adds_line.replace("q", "")

        path = str(pathlib.Path(__file__).parent.resolve()) + "/.." + adds_line + "data/state_data.txt"

        txt_data = open(path,"w")
        sv_data = ""
        for i in range(0, len(self.BuildingState)):
            if i != 0:
                sv_data = sv_data + ";" + self.BuildingState[i]
            else:
                sv_data = self.BuildingState[i]

        txt_data.write(sv_data)
        txt_data.close()

        return self.BuildingState